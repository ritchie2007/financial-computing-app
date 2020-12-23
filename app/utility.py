from random import random
import time
from time import strftime, localtime
from datetime import timedelta, datetime #, date
from pytz import timezone
from dateutil.relativedelta import relativedelta
from flask_login import current_user, login_user, logout_user, login_required
from flask import render_template, flash, redirect, request, url_for, session, make_response, json
from flask import send_from_directory, send_file, after_this_request
from sqlalchemy import desc, asc, func, or_ # for table.order_by(Task.enddate).all()
from app import app, db
from app.forms import LoginForm, RegistrationForm, EditProfileForm, WebNavForm
from app.models import User, Data_table, activity_code, CorporationReport, Staff, Task, \
    Timesheet, Corporation, Individual, Userlog, Mulform, TimesheetTempData

import os # 目录及文件操作
import openpyxl
from openpyxl import Workbook, load_workbook # pip install openpyxl 
from openpyxl.styles import Font, Color, colors, PatternFill, Border, borders, fills, Fill, alignment, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, Series, LineChart, ScatterChart

def month_offset(start_date, number_of_month):
    return datetime.strptime(start_date, '%Y-%m-%d') + relativedelta(months=number_of_month)
# print(get_today_month(-3))
print(month_offset('2019-01-31', 1))

def get_id_from_name(name, start, count):
    id_list = []
    for i in range(count):
        item = ((name[i+start].split(" | "))[0].lstrip(' ').lstrip('0'))
        if item not in id_list and item != "" :
            id_list.append(item)
    # for i in range(count - len(id_list)):
    #     id_list.append("")
    return(id_list)

# contact
# 在个人里添加(修改删除)公司后，引起公司表发生相应动作
def indiv_to_corp_contact(value1, corpid, oper, value0): # 人里公司改，公司里人改
    print(' --- contact ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Corporation.query.get(x)
                    tmp = (my_data.contact)
                    if len(tmp) > 0:
                        tmp = (my_data.contact).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.contact = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Corporation.query.get(x)
            if my_data.contact == "":
                my_data.contact = str(corpid)
            else:
                print(" -- my_data -- " + my_data.contact)
                tmp = (my_data.contact).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.contact = ",".join(tmp)
            print(my_data.contact)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Corporation.query.get(x)
                tmp = (my_data.contact)
                if len(tmp) >0:
                    tmp = (my_data.contact).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.contact = ",".join(tmp)
                        db.session.commit()
        # flash("Corp from Individual was updated in Corp table")
# 在公司里添加(修改删除)个人后，引起个人表发生相应动作
def corp_contact_to_indiv(value1, corpid, oper, value0): # 公司里人改，人里公司改
    print(' --- utility contact ---')
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Individual.query.get(x)
                    tmp = (my_data.contact_corp)
                    if len(tmp) > 0:
                        tmp = (my_data.contact_corp).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.contact_corp = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Individual.query.get(x)
            if my_data.contact_corp == "":
                my_data.contact_corp = str(corpid)
            else:
                print(" -- my_data -- " + my_data.contact_corp)
                tmp = (my_data.contact_corp).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.contact_corp = ",".join(tmp)
            print(my_data.contact_corp)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Individual.query.get(x)
                tmp = (my_data.contact_corp)
                if len(tmp) >0:
                    tmp = (my_data.contact_corp).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.contact_corp = ",".join(tmp)
                        db.session.commit()

    # flash("Contact from Corp was updated in Individual table")

# director
# 在个人里添加(修改删除)公司后，引起公司表发生相应动作
def indiv_to_corp_director(value1, corpid, oper, value0): # 人里公司改，公司里人改
    print(' --- director ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Corporation.query.get(x)
                    tmp = (my_data.director)
                    if len(tmp) > 0:
                        tmp = (my_data.director).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.director = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Corporation.query.get(x)
            if my_data.director == "":
                my_data.director = str(corpid)
            else:
                print(" -- my_data -- " + my_data.director)
                tmp = (my_data.director).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.director = ",".join(tmp)
            print(my_data.director)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Corporation.query.get(x)
                tmp = (my_data.director)
                if len(tmp) >0:
                    tmp = (my_data.director).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.director = ",".join(tmp)
                        db.session.commit()
        # flash("Director from Individual was updated in Corp table")
# 在公司里添加(修改删除)个人后，引起个人表发生相应动作
def corp_director_to_indiv(value1, corpid, oper, value0): # 公司里人改，人里公司改
    print(' --- director ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Individual.query.get(x)
                    tmp = (my_data.director_corp)
                    if len(tmp) > 0:
                        tmp = (my_data.director_corp).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.director_corp = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Individual.query.get(x)
            if my_data.director_corp == "":
                my_data.director_corp = str(corpid)
            else:
                print(" -- my_data -- " + my_data.director_corp)
                tmp = (my_data.director_corp).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.director_corp = ",".join(tmp)
            print(my_data.director_corp)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Individual.query.get(x)
                tmp = (my_data.director_corp)
                if len(tmp) >0:
                    tmp = (my_data.director_corp).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.director_corp = ",".join(tmp)
                        db.session.commit()

    # flash("Director from Corp was updated in Individual table")

# shareholder(个人作为shareholder)
# 在个人里添加(修改删除)公司后，引起公司表发生相应动作
def indiv_to_corp_shareholder(value1, corpid, oper, value0): # 人里公司改，公司里人改
    print(' --- shareholder ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Corporation.query.get(x)
                    tmp = (my_data.shareholder)
                    if len(tmp) > 0:
                        tmp = (my_data.shareholder).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.shareholder = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Corporation.query.get(x)
            if my_data.shareholder == "":
                my_data.shareholder = str(corpid)
            else:
                print(" -- my_data -- " + my_data.shareholder)
                tmp = (my_data.shareholder).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.shareholder = ",".join(tmp)
            print(my_data.shareholder)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Corporation.query.get(x)
                tmp = (my_data.shareholder)
                if len(tmp) >0:
                    tmp = (my_data.shareholder).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.shareholder = ",".join(tmp)
                        db.session.commit()
    # flash("Shareholder from Indiv was updated in Corp table")
# 在公司里添加(修改删除)个人后，引起个人表发生相应动作
# corp_shareholder_to_indiv在公司里添加(修改删除)个人作为shareholder后，引起indiv表发生相应动作
def corp_shareholder_to_indiv(value1, corpid, oper, value0): # 公司里人改，人里公司改
    # value1是从前台获取的要添加的个人信息，corpid是即将生成的公司的id，oper中0表示添加，1表示修改，2表示删除
    # value0是原来的信息(比如原来shareholder中个人id)，因为是插入所以value0=''
    print(' --- shareholder ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0: # 如果原来shareholder有值，比如说有2个sharehold，就是说有2个个人的id
            for x in value0:
                if x not in value1: # 如果原来shareholder个人id不在新的名单里，表示原来的id被删除了
                    my_data = Individual.query.get(x) # 从个人表里 获取信息
                    tmp = (my_data.sharehold_corp) # 获取该个人里面的 sharehold公司信息
                    if len(tmp) > 0: # 如果有
                        tmp = (my_data.sharehold_corp).split(',') # 形成数组
                        if str(corpid) in tmp: # 找到对应的公司，予以删除
                            tmp.remove(str(corpid))
                            my_data.sharehold_corp = ",".join(tmp)
                            db.session.commit()

        for x in value1: # 新的sharehold信息
            my_data = Individual.query.get(x) # 找到对应的个人记录
            if my_data.sharehold_corp == "":# 如果个人记录中 sharehold部分没有内容，则直接添加
                my_data.sharehold_corp = str(corpid)
            else:
                print(" -- my_data -- " + my_data.sharehold_corp)
                tmp = (my_data.sharehold_corp).split(',') # 如果个人记录中 sharehold有内容，则只添加新的
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.sharehold_corp = ",".join(tmp)
            print(my_data.sharehold_corp)
            db.session.commit()

    elif oper == 2: # 删除，找到个人记录，逐个删除sharehold部分 对应的 该公司 内容
        if len(value0) > 0:
            for x in value0:
                my_data = Individual.query.get(x)
                tmp = (my_data.sharehold_corp)
                if len(tmp) >0:
                    tmp = (my_data.sharehold_corp).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.sharehold_corp = ",".join(tmp)
                        db.session.commit()
    # flash("Shareholder from Corp was updated in Individual table")

# shareholder(公司作为shareholder)
# 在公司里添加(修改删除)作为shareholders的公司后，引起作为sharehold的相应公司表发生相应动作
# 不能添加 目前的公司 （corp, as shareholder for) 作为其他公司的shareholder，因为你并不知道是作为那个公司
# 的第几个shareholder
# 公司里shareholder_corp变，引起对应公司里面corp_as_shareholder改变
def corp_shareholder_to_corp(value1, corpid, oper, value0): # 公司里sharehold公司改，相应sharehold公司里公司改
    print(' --- shareholder ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Corporation.query.get(x)
                    tmp = (my_data.corp_as_shareholder)
                    if tmp:
                        if len(tmp) > 0:
                            tmp = (my_data.corp_as_shareholder).split(',')
                            if str(corpid) in tmp:
                                tmp.remove(str(corpid))
                                my_data.corp_as_shareholder = ",".join(tmp)
                                db.session.commit()

        for x in value1:
            my_data = Corporation.query.get(x)
            if my_data.corp_as_shareholder is None:
                my_data.corp_as_shareholder = str(corpid)
            elif my_data.corp_as_shareholder == "":
                my_data.corp_as_shareholder = str(corpid)
            else:
                tmp = (my_data.corp_as_shareholder).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.corp_as_shareholder = ",".join(tmp)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Corporation.query.get(x)
                if my_data.corp_as_shareholder is not None:
                    tmp = (my_data.corp_as_shareholder)
                if len(tmp) >0:
                    tmp = (my_data.corp_as_shareholder).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.corp_as_shareholder = ",".join(tmp)
                        db.session.commit()
    # flash("updated in Corp table")
# corp_as_shareholder改变（比如删除该公司了），引起对应的公司里shareholder_corp变
def corp_to_corp_shareholder(value1, corpid, oper, value0): 
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Corporation.query.get(x)
                    tmp = (my_data.shareholder_corp)
                    if tmp:
                        if len(tmp) > 0:
                            tmp = (my_data.shareholder_corp).split(',')
                            if str(corpid) in tmp:
                                tmp.remove(str(corpid))
                                my_data.shareholder_corp = ",".join(tmp)
                                db.session.commit()

        for x in value1:
            my_data = Corporation.query.get(x)
            if my_data.shareholder_corp is None:
                my_data.shareholder_corp = str(corpid)
            elif my_data.shareholder_corp == "":
                my_data.shareholder_corp = str(corpid)
            else:
                tmp = (my_data.shareholder_corp).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.shareholder_corp = ",".join(tmp)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Corporation.query.get(x)
                if my_data.shareholder_corp is not None:
                    tmp = (my_data.shareholder_corp)
                if len(tmp) >0:
                    tmp = (my_data.shareholder_corp).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.shareholder_corp = ",".join(tmp)
                        db.session.commit()

# spouse
def indiv_to_spouse(value1, corpid, oper, value0): # 配偶改
    print(' --- spouse ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Individual.query.get(x)
                    tmp = (my_data.spouse)
                    if len(tmp) > 0:
                        tmp = (my_data.spouse).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.spouse = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Individual.query.get(x)
            if my_data.spouse == "":
                my_data.spouse = str(corpid)
            else:
                print(" -- my_data -- " + my_data.spouse)
                tmp = (my_data.spouse).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.spouse = ",".join(tmp)
            print(my_data.spouse)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Individual.query.get(x)
                tmp = (my_data.spouse)
                if len(tmp) >0:
                    tmp = (my_data.spouse).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.spouse = ",".join(tmp)
                        db.session.commit()
    # flash("Spouse Updated Successfully")

# parents
def parent_to_child(value1, corpid, oper, value0): # 父里子改，子里父改
    print(' --- parents ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Individual.query.get(x)
                    tmp = (my_data.child)
                    if len(tmp) > 0:
                        tmp = (my_data.child).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.child = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Individual.query.get(x)
            if my_data.child == "":
                my_data.child = str(corpid)
            else:
                print(" -- my_data -- " + my_data.child)
                tmp = (my_data.child).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.child = ",".join(tmp)
            print(my_data.child)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Individual.query.get(x)
                tmp = (my_data.child)
                if len(tmp) >0:
                    tmp = (my_data.child).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.child = ",".join(tmp)
                        db.session.commit()
    # flash("parents updated")

# child
def child_to_parent(value1, corpid, oper, value0): # 子里父改，父里子改
    print(' --- child ---')
    print(value1)
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Individual.query.get(x)
                    tmp = (my_data.parent)
                    if len(tmp) > 0:
                        tmp = (my_data.parent).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.parent = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Individual.query.get(x)
            if my_data.parent == "":
                my_data.parent = str(corpid)
            else:
                print(" -- my_data -- " + my_data.parent)
                tmp = (my_data.parent).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.parent = ",".join(tmp)
            print(my_data.parent)
            db.session.commit()

    elif oper == 2:
        if len(value0) > 0:
            for x in value0:
                my_data = Individual.query.get(x)
                tmp = (my_data.parent)
                if len(tmp) >0:
                    tmp = (my_data.parent).split(',')
                    if str(corpid) in tmp:
                        tmp.remove(str(corpid))
                        my_data.parent = ",".join(tmp)
                        db.session.commit()
    # flash("Child Updated Successfully")

# get Index array
def get_index_index(type, id_list, dropdown_list): # 由ID字串获取下拉菜单index字串
    index_list = []
    tmp0 = []
    for x in dropdown_list:
        if type == 'Corporation':
            tmp0.append(str(x.indiv_id))
        elif type == 'Individual':
            tmp0.append(str(x.corp_id))

    for x in id_list:
        tmp1 = []
        if x:
            if (len(x)>0):
                x = x.split(',')
                for y in x:
                    if y:
                        if len(y)>0:
                            if tmp0.count(y)>0:
                                tmp1.append(tmp0.index(y))
        index_list.append(tmp1)
    return index_list

# convert str to int (index)
def convert_to_int(data_str):
    data_int = 0
    data_str = data_str.replace(' ', '')
    if data_str.isdigit():
        data_int = int(data_str)
    return data_int

# dailyentry to report
def dailyentry_to_report(value1, corpid, oper, value0): # 公司里人改，人里公司改
    if value0 == "":
        value0 = []
    else:
        value0 = value0.split(',')

    if oper < 2:
        if len(value0) > 0:
            for x in value0:
                if x not in value1:
                    my_data = Individual.query.get(x)
                    tmp = (my_data.director_corp)
                    if len(tmp) > 0:
                        tmp = (my_data.director_corp).split(',')
                        if str(corpid) in tmp:
                            tmp.remove(str(corpid))
                            my_data.director_corp = ",".join(tmp)
                            db.session.commit()

        for x in value1:
            my_data = Individual.query.get(x)
            if my_data.director_corp == "":
                my_data.director_corp = str(corpid)
            else:
                tmp = (my_data.director_corp).split(',')
                if (str(corpid) not in tmp):
                    tmp.append(str(corpid))
                    my_data.director_corp = ",".join(tmp)
            db.session.commit()

# excel report
# download forntend
def download_munu(category):
    filtertxt = ['','','']
    filterdata = ['','','']
    if category == 'Corporation':
        filtertxt = ['','','']
        filterdata = ['','','']
    elif category == 'Individual':
        filtertxt = ['','','']
        filterdata = ['','','']
    elif category == 'Task':
        filtertxt = ['Period_end from','Period_end to','']
        filterdata[0] = db.session.query(Task.periodend.label('startdate')).distinct().order_by(Task.periodend.desc()).all()
        filterdata[1] = db.session.query(Task.periodend.label('startdate')).distinct().order_by(Task.periodend.desc()).all()
        filterdata[2] = ''
    elif category == 'Timesheet':
        filtertxt = ['Date from','Date to','Staff name']
        filterdata[0] = db.session.query(Timesheet.startdate).distinct().order_by(Timesheet.startdate.desc()).all()
        filterdata[1] = db.session.query(Timesheet.startdate).distinct().order_by(Timesheet.startdate.desc()).all()
        filterdata[2] = db.session.query(Timesheet.staff.label('name')).distinct().order_by(Timesheet.staff.asc()).all()
    elif category == 'Staff':
        filtertxt = ['Date from','Date to','Staff name']
        filterdata[0] = db.session.query(Staff.startdate).distinct().order_by(Staff.startdate.desc()).all()
        filterdata[1] = db.session.query(Staff.startdate).distinct().order_by(Staff.startdate.desc()).all()
        filterdata[2] = db.session.query(Staff.name).distinct().order_by(Staff.name.asc()).all()
    elif category == 'Corporation report':
        filtertxt = ['Date from','Date to','Corporation name']
        filterdata[0] = db.session.query(CorporationReport.startdate).distinct().order_by(CorporationReport.startdate.desc()).all()
        filterdata[1] = db.session.query(CorporationReport.startdate).distinct().order_by(CorporationReport.startdate.desc()).all()
        filterdata[2] = db.session.query(CorporationReport.corp.label('name')).distinct().order_by(CorporationReport.corp.asc()).all()
    return (filtertxt, filterdata)

def table_header(idx):
    head = ['','','','','','']
    col = ['','','','','','']
    note = ['','','','','','']
    head[0] = ["Corporation ID", "Bussiness No.", "Corporation Name", "Registration Name", "M Business Name", "M Licenses period", "Incorporation Date", "Type of Incorporation", "Anniversary Date", "Corporation Key", "Corporation No.", "Ontario Corp No.", "Provincial No. 1", "Provincial No. 2", "Business Address", "Mailing Address", "CRA Contact Person", "CRA Contact Phone", "Office Fax No.", "Email", "Website", "Wechat", "Tax Year End", "HST Reporting", "HST Report_method ", "HST Account No.", "HST Account", "Payroll  Account No.", "Payroll_T4", "Subcontractor_T4A", "Dividend_T5", "Withhold_Tax No.", "Withhold_Tax_account ", "CRA Other Account No.", "CRA Other Account", "WSIB Account No.", "WSIB Account", "EHT Account No.", "EHT Account", "Other Account No.", "Other Account", "Holding Corporation ", "Ovesea Invest Corp ", "Corporation Indursty", "Corporation Info", "Specific Info", "Bank Account 1", "Bank Account 2", "Bank Account 3", "Accounting Software", "Software Password", "Payroll Software", "Payroll Software password", "Accounting Service ", "Accounting Service Type", "Leading ", "contact id", "Director id", "Shareholder id"]
    # len(head[0]), len(col[0]), len(note[0]) = 59 62 3
    col[0] = ["corp_id", "corp1", "corp2", "corp3", "corp4", "corp5", "corp6", "corp7", "corp8", "corp9", "corp10", "corp11", "corp12", "corp13", "corp14", "corp15", "corp16", "corp17", "corp18", "corp19", "corp20", "corp21", "corp22", "corp23", "corp24", "corp25", "corp26", "corp27", "corp28", "corp29", "corp30", "corp31", "corp32", "corp33", "corp34", "corp35", "corp36", "corp37", "corp38", "corp39", "corp40", "corp41", "corp42", "corp43", "corp44", "corp45", "corp46", "corp47", "corp48", "corp49", "corp50", "corp51", "corp52", "corp53", "corp54", "corp55", "corp56", "corp57", "corp58", "contact", "director", "shareholder"]
    note[0] = [["corp5", "corp6"],["corp8", "corp9"], ["corp18", "corp19"]]
    head[1] = ["Individual ID", "SIN", "Name", "other_name", "email", "phone1", "phone2", "address1", "address2", "mail_address", "wechat", "Sole_proprietor", "HST_report", "Payroll", "Withhold_tax", "WSIB", "CRA_other", "Oversea_asset_t1135", "Oversea_corp_t1134", "Tslip", "Tax_personal_info", "Specific_info", "Engage_account", "Engage_leading", "note", "Corporation(Contact of) ", "Corporation(Director of)", "Corporation(Shareholder of)", "Spouse", "Parent", "Children"]
    # len(head[1]), len(col[1]), len(note[1]) = 31 33 3
    col[1] = ["indiv_id", "sin", "prefix", "last_name", "first_name", "other_name", "email", "phone1", "phone2", "address1", "address2", "mail_address", "wechat", "cra_sole_proprietor", "cra_hst_report", "cra_payroll", "cra_withhold_tax", "cra_wsib", "cra_other", "oversea_asset_t1135", "oversea_corp_t1134", "tslip", "tax_personal_info", "specific_info", "engage_account", "engage_leading", "note", "contact_corp", "director_corp", "sharehold_corp", "spouse", "parent", "child"]
    note[1] = ["prefix", "last_name", "first_name"]
    head[2] = ["Task_id", "Client_Corp_id", "Client_Corp_name", "Corp_Bussi_No.", "Jobtype", "Period end", "Responsible Person", "Start date", "End date", "Status", "Details", "Recurrence", "Priority", "Work hour"]
    # len(head[2]), len(col[2]), len(note[2]) = 14 14 0
    col[2] = ["task_id", "client_corp_id", "client_corp_name", "client_corp_bussi_no", "jobtype_code", "periodend", "responsible", "startdate", "enddate", "status", "details", "recurrence", "priority", "worktime"]
    note[2] = []
    head[3] = ["Timesheet_id", "Startdate", "Calendar hour", "Adjust hour", "Work hour", "Timesheet name", "Timesheet content", "Activity_type", "Corporation 1", "Corporation 2", "Corporation 3", "Corporation 4", "Staff", "Average time"]
    # len(head[3]), len(col[3]), len(note[3]) = 14 15 2
    col[3] = ["timesheet_id", "startdate", "calhour", "adjhour", "adjmin", "workhour", "entryname", "entrycontent", "activitytype", "corp1", "corp2", "corp3", "corp4", "staff", "avgtime"]
    note[3] = ["adjhour", "adjmin"]
    head[4] = ["Staff_id", "Name", "Startdate", "Job/Task", "Calendar hour", "Adjust hour", "Work hour"]
    # len(head[4]), len(col[4]), len(note[4]) = 7 8 2
    col[4] = ["staff_id", "name", "startdate", "job", "calendarhour", "adjhour", "adjmin", "workhour"]
    note[4] = ["adjhour", "adjmin"]
    head[5] = ["Report_id", "Corporation", "Start date", "Timesheet name", "Activity_type", "Timesheet content", "Work hour", "Job_id"]
    # len(head[5]), len(col[5]), len(note[5]) = 8 8 0
    col[5] = ["corp_report_id", "corp", "startdate", "entryname", "activitytype", "entrycontent", "workhour", "jobid"]
    note[5] = []
    return (head[idx], col[idx], note[idx])

def get_data(idx, filters):    
    if idx == 0:
        list_data = db.session.query(Corporation).order_by(Corporation.corp_id.asc()).all()
    elif idx == 1:
        list_data = db.session.query(Individual).order_by(Individual.indiv_id.asc()).all()
    elif idx == 2:
        attribute = '' if filters[0] == '' else 'Task.periodend >= filters[0]'
        attribute = attribute if filters[1] == '' else attribute + ', Task.periodend <= filters[1]'
        attribute = attribute if filters[2] == '' else attribute + ', Task.client_corp_name == filters[2]'
        att = []
        exec('att.append(db.session.query(Task).filter({}).order_by(Task.periodend.desc()).all())'.format(attribute.lstrip(', ')))
        list_data = att[0]
    elif idx == 3:
        attribute = '' if filters[0] == '' else 'Timesheet.startdate >= filters[0]'
        attribute = attribute if filters[1] == '' else attribute + ', Timesheet.startdate <= filters[1]'
        attribute = attribute if filters[2] == '' else attribute + ', Timesheet.staff == filters[2]'
        att = []
        exec('att.append(db.session.query(Timesheet).filter({}).order_by(Timesheet.startdate.desc()).all())'.format(attribute.lstrip(', ')))
        list_data = att[0]
    elif idx == 4: # staff
        attribute = '' if filters[0] == '' else 'Staff.startdate >= filters[0]'
        attribute = attribute if filters[1] == '' else attribute + ', Staff.startdate <= filters[1]'
        attribute = attribute if filters[2] == '' else attribute + ', Staff.name == filters[2]'
        att = []
        exec('att.append(db.session.query(Staff).filter({}).order_by(Staff.startdate.desc()).all())'.format(attribute.lstrip(', ')))
        list_data = att[0]
    elif idx == 5: # corp report
        attribute = '' if filters[0] == '' else 'CorporationReport.startdate >= filters[0]'
        attribute = attribute if filters[1] == '' else attribute + ', CorporationReport.startdate <= filters[1]'
        attribute = attribute if filters[2] == '' else attribute + ', CorporationReport.corp == filters[2]'
        att = []
        exec('att.append(db.session.query(CorporationReport).filter({}).order_by(CorporationReport.startdate.desc()).all())'.format(attribute.lstrip(', ')))
        list_data = att[0]

        # if filters[0] != '' and filters[1] == '' and filters[2] == '':
        #     list_data = CorporationReport.query.filter(CorporationReport.periodend >= filters[0]).all()
        # elif filters[0] != '' and filters[1] != '' and filters[2] == '':
        #     list_data = CorporationReport.query.filter(CorporationReport.periodend >= filters[0], CorporationReport.periodend <= filters[1]).all()
        # elif filters[0] != '' and filters[1] == '' and filters[2] != '':
        #     list_data = CorporationReport.query.filter(CorporationReport.periodend >= filters[0], CorporationReport.corp == filters[2]).all()
        # elif filters[0] != '' and filters[1] != '' and filters[2] != '':
        #     list_data = CorporationReport.query.filter(CorporationReport.periodend >= filters[0], CorporationReport.periodend <= filters[1], CorporationReport.corp == filters[2]).all()
        # elif filters[0] == '' and filters[1] != '' and filters[2] == '':
        #     list_data = CorporationReport.query.filter(CorporationReport.periodend <= filters[1]).all()
        # elif filters[0] == '' and filters[1] != '' and filters[2] != '':
        #     list_data = CorporationReport.query.filter(CorporationReport.periodend <= filters[1], CorporationReport.CorporationReport == filters[2]).all()
        # elif filters[0] == '' and filters[1] == '' and filters[2] != '':
        #     list_data = CorporationReport.query.filter(CorporationReport.corp == filters[2]).all()
    return list_data

def excel_export(cat, filters, fname):
    fname = fname
    idx = ['Corporation','Individual','Task','Timesheet','Staff','Corporation report'].index(cat)
    print(idx)
    pram = table_header(idx)
    wb = Workbook()
    ws = wb.active
    ws.title = cat # sheet name
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(color="FF0000", bold=True, name='Arial', size=15)
    ws['A1'] = fname
    ws.merge_cells('A2:J2')
    ws['A2'].font = Font(color="0000FF", bold=True, name='Arial', size=10)
    if idx == 2:
        ws['A2'] = 'Start Date: ' + filters[0] + ', End Date: ' + filters[1]
    elif idx == 3 or idx == 4:
        ws['A2'] = 'Start Date: ' + filters[0] + ', End Date: ' + filters[1] + ', Staff name: ' + filters[2]
    elif idx == 5:
        ws['A2'] = 'Start Date: ' + filters[0] + ', End Date: ' + filters[1] + ', Corporation name: ' + filters[2]
    ws.row_dimensions[3].fill = PatternFill("solid", fgColor="DDDDDD")
    ws['A3'] = ''
    ws['A3'].fill = PatternFill("solid", fgColor="DDDDDD")
    ws.append(pram[0])
    cols = len(pram[0])
    for i in range(cols):
        ws.cell(row=4, column = i+1).font = Font(color="000000", bold=True, name='Arial', size=10)
    
    my_data = get_data(idx, filters)
    if idx == 0:
        for row in my_data:
            rdata = []
            for i in range(5):
                exec('rdata.append(row.{})'.format(pram[1][i]))
            rdata.append(row.corp5 + ' - ' + row.corp6)
            rdata.append(row.corp7)
            rdata.append(row.corp8)
            rdata.append(row.corp9 + ' - ' + row.corp10)
            for i in range(7):
                exec('rdata.append(row.{})'.format(pram[1][i+11]))
                print('corporation-', pram[1][i])
            rdata.append(row.corp19 + ' - ' + row.corp18)
            for i in range(42):
                exec('rdata.append(row.{})'.format(pram[1][i+20]))
            ws.append(rdata)
    elif idx == 1:
        for row in my_data:
            rdata = []
            rdata.append(row.indiv_id)
            rdata.append(row.sin)
            rdata.append(row.prefix + ' ' + row.last_name + ', ' + row.first_name)
            for i in range(28):
                exec('rdata.append(row.{})'.format(pram[1][i+5]))
            ws.append(rdata)
    elif idx == 2:
        for row in my_data:
            rdata = []
            for i in range(14):
                exec('rdata.append(row.{})'.format(pram[1][i]))
            ws.append(rdata)
    elif idx == 3:
        for row in my_data:
            rdata = []
            rdata.append(row.timesheet_id)
            rdata.append(row.startdate)
            rdata.append(row.calhour)
            rdata.append(row.adjhour + ' : ' + row.adjmin)
            for i in range(10):
                exec('rdata.append(row.{})'.format(pram[1][i+5]))
            ws.append(rdata)
    elif idx == 4:
        for row in my_data:
            rdata = []
            for i in range(5):
                exec('rdata.append(row.{})'.format(pram[1][i]))
            rdata.append(row.adjhour + ' : ' + row.adjmin)
            rdata.append(row.workhour)
            ws.append(rdata)
    elif idx == 5:
        for row in my_data:
            rdata = []
            for i in range(8):
                exec('rdata.append(row.{})'.format(pram[1][i]))
            ws.append(rdata)
    print(fname)
    wb.save(app.config["CLIENT_CSV"] + fname)

# user records
def userrecrods(user, field):
    '''records of user'''
    username = user
    email = ''
    password = ''
    ip = ''
    datadate = datetime.now(timezone('America/Toronto')).strftime("%Y-%m-%d %H:%M:%S")
    datatime = int(time.time())
    badfield = field
    if field == '':
        status = 'True'
        hourlock = 0
        daylock = 0
        attemptafterlock = 0
    else:
        status = 'False'
        count = db.session.query(Userlog).with_entities(func.count(Userlog.log_id)).filter(Userlog.datatime >= (datatime - 3600), Userlog.username == user, Userlog.status == 'False').scalar()
        print('hourlock count', count)
        hourlock = datatime if count > 1 else 0
        count = db.session.query(Userlog).with_entities(func.count(Userlog.log_id)).filter(Userlog.datatime >= (datatime - 10800), Userlog.username == user, Userlog.status == 'False').scalar()
        print('daylock count', count)
        daylock = datatime if count > 4 else 0
        attemptafterlock = count + 1
    # attemptafterlock = 0
    my_data = Userlog(username, email, password, ip, datadate, datatime, badfield, hourlock, daylock, status, attemptafterlock)
    db.session.add(my_data)
    db.session.commit()
    return status

# authentication
def authentication(user):
    '''
    Login Input --> Registered User? --> Y, expired? --Y--> authentication False return()
                                                        No --> check log
                    Not user --> check log 
                    check log: --locked--> authentication False
    '''
    authentication = False
    au = db.session.query(User).filter(User.username == user).scalar()
    currenttime = int(time.time())
    print('au: ', au, '  currenttime: ', currenttime)
    if au:
        print(' ---- registered user ---')
        if au.identification is None or au.identification == '' or au.identification == 0:
            print(' ---- identification: None, empty or 0 --- ', au.identification)
            authentication = True
        elif currenttime > int(au.identification):
            print('--- > identification false---')
            return authentication
        else:
            authentication = True
    
    log_id = db.session.query(func.max(Userlog.log_id)).filter(Userlog.username == user).scalar()
    if log_id is None:
        print('log_id: ', log_id)
        authentication = True
    else:
        my_data = Userlog.query.get(log_id)
        print('log_id: ', log_id, ' mydata: ', my_data)
        if (my_data.hourlock == 0 and my_data.daylock == 0):
            print('---my_data.hourlock == 0 and my_data.daylock == 0---')
            authentication = True
        elif (currenttime > (my_data.hourlock + 3600)) and (currenttime > (my_data.daylock + 86400)):
            authentication = True
            my_data.hourlock = 0
            my_data.daylock = 0
            db.session.add(my_data)
            db.session.commit()
            print('---- over the limit time -------')
        else:
            authentication = False

    return authentication

# import excel file
def import_excel_corp():
    workbook = openpyxl.load_workbook("/Users/Ritchie/Downloads/export from CCH_20201029_Updated.xlsx")
    worksheet = workbook.get_sheet_by_name('corporation2')
    # row3=[item.value for item in list(worksheet.rows)[2]] # print('-- entire 3rd row data --',row3)
    # col3=[item.value for item in list(worksheet.columns)[2]]  # print('-- entire 3rd column data --',col3)
    # cell_2_3 = worksheet.cell(row = 2,column = 3).value
    # print('------',cell_2_3)
    # cell_2_3 = worksheet.cell(row = 312,column = 4).value
    # print('------',cell_2_3)
    # max_row = worksheet.max_row
    # print('max ',max_row)
    for i in range(3, 312):
        r = []
        for x in range (1,37):
            da = worksheet.cell(row = i, column = x).value
            if da:
                r.append(da)
            else:
                r.append('')
        print(r)
        corp1 = r[1]
        corp2 = r[0]
        corp3 = r[2]
        corp4 = r[4]
        corp5 = r[5]
        corp6 = r[6]
        corp7 = r[7]
        corp8 = r[8]
        anniv = r[9]
        if anniv:
            anniv = anniv.lower().replace('to','').replace(' ','').replace('-','')
            corp9 = anniv[0:2] + '/' + anniv[2:4]
            corp10 = anniv[4:6] + '/' + anniv[6:8]
        else:
            corp9 = ''
            corp10 = ''
        corp11 = r[10]
        corp12 = r[3]
        corp13 = r[11]
        corp14 = r[12]
        corp15 = ''
        if r[13]:
            corp16 = 'Contact Group: ' + r[13]
        corp17 = ''
        for x in range(15,21):
            if r[x-1]:
               corp17 = corp17  + ' ' + r[x-1]

        corp18 = ''
        corp19 = r[23]
        corp20 = r[24]
        corp21 = ''
        corp22 = r[21]
        corp23 = ''
        corp24 = ''
        corp25 = r[22]
        corp26 = r[32]
        corp27 = r[31]
        corp28 = ''
        corp29 = ''
        corp30 = ''
        corp31 = ''
        corp32 = ''
        corp33 = ''
        corp34 = ''
        corp35 = ''
        corp36 = ''
        corp37 = ''
        corp38 = ''
        corp39 = ''
        corp40 = ''
        corp41 = ''
        corp42 = ''
        corp43 = ''
        corp44 = ''
        corp45 = ''
        corp46 = ''
        corp47 = r[27]
        corp48 = r[28]
        corp49 = ''
        corp50 = ''
        corp51 = ''
        corp52 = r[25]
        corp53 = ''
        corp54 = ''
        corp55 = ''
        corp56 = r[26]
        corp57 = ''
        corp58 = ''

        if corp20:
            corp201 = corp20.replace('-','')
        else:
            corp201 = ''
            
        if corp21:
            corp211 = corp21.replace('-','')
        else:
            corp211 = ''

        task = 0
        recent_update = ""
        timemark = datetime.utcnow()
        contact_position = ""
        shareholder_info = ""
        shareholder_corp_info = ""
        corp_as_shareholder = ""
        contact = ""
        director = ""
        shareholder = ""
        shareholder_corp = ""
        # print(corp1, corp2, corp3, corp4, corp5, corp6, corp7, corp8, corp9, corp10, corp11, corp12, corp13, corp14, corp15, corp16, corp17, corp18, corp19, corp20, corp21, corp201, corp211, corp22, corp23, corp24, corp25, corp26, corp27, corp28, corp29, corp30, corp31, corp32, corp33, corp34, corp35, corp36, corp37, corp38, corp39, corp40, corp41, corp42, corp43, corp44, corp45, corp46, corp47, corp48, corp49, corp50, corp51, corp52, corp53, corp54, corp55, corp56, corp57, corp58, contact, director, shareholder, task, recent_update, contact_position, shareholder_info, shareholder_corp, shareholder_corp_info, corp_as_shareholder, timemark)
        my_data = Corporation(corp1, corp2, corp3, corp4, corp5, corp6, corp7, corp8, corp9, corp10, corp11, corp12, corp13, corp14, corp15, corp16, corp17, corp18, corp19, corp20, corp21, corp201, corp211, corp22, corp23, corp24, corp25, corp26, corp27, corp28, corp29, corp30, corp31, corp32, corp33, corp34, corp35, corp36, corp37, corp38, corp39, corp40, corp41, corp42, corp43, corp44, corp45, corp46, corp47, corp48, corp49, corp50, corp51, corp52, corp53, corp54, corp55, corp56, corp57, corp58, contact, director, shareholder, task, recent_update, contact_position, shareholder_info, shareholder_corp, shareholder_corp_info, corp_as_shareholder, timemark)
        db.session.add(my_data)
        db.session.commit()
    flash("Import Successfully")
    msg = 'successful'
    return msg

# import excel file
def import_excel_indiv():
    workbook = openpyxl.load_workbook("/Users/Ritchie/Downloads/export from CCH_20201029_Updated.xlsx")
    worksheet = workbook.get_sheet_by_name('individual')
    # row3=[item.value for item in list(worksheet.rows)[2]] # print('-- entire 3rd row data --',row3)
    # col3=[item.value for item in list(worksheet.columns)[2]]  # print('-- entire 3rd column data --',col3)
    # cell_2_3 = worksheet.cell(row = 2,column = 3).value
    # print('------',cell_2_3)
    # cell_2_3 = worksheet.cell(row = 312,column = 4).value
    # print('------',cell_2_3)
    # max_row = worksheet.max_row
    # print('max ',max_row)
    for i in range(3, 372):
        r = []
        for x in range (1,16):
            da = worksheet.cell(row = i, column = x).value
            if da:
                r.append(da)
            else:
                r.append('')
        print(r)
        sin = r[13]
        prefix = ''
        last_name = r[2]
        first_name = r[1]
        other_name = ''
        email = r[5]
        phone1 = '(h)' + r[3] + '(w)' + r[14]
        phone2 = r[4]
        if phone1:
            phone1digit = phone1.replace('-','')
        else:
            phone1digit = ''
        if phone2:
            phone2digit = phone2.replace('-','')
        else:
            phone2digit = ''
        address1 = r[6] + '-' + r[7] + ' ' + r[8] + ', ' + r[9] + ', ' + r[10] + ' ' + r[11] + ', ' + r[12]
        address2 = ''
        mail_address = '' 
        wechat = ''
        cra_sole_proprietor = ''
        cra_hst_report = ''
        cra_payroll = ''
        cra_withhold_tax = ''
        cra_wsib = ''
        cra_other = ''
        oversea_asset_t1135 = ''
        oversea_corp_t1134 = ''
        tslip = ''
        tax_personal_info = ''
        specific_info = ''
        engage_account = ''
        engage_leading = ''
        note = ''
        timemark = datetime.utcnow()
        contact_corp = ""
        director_corp = ""
        sharehold_corp = ""
        spouse = ""
        parent = ""
        child = ""
        
        my_data = Individual(sin, prefix, last_name, first_name, other_name, email, phone1, phone2, phone1digit, phone2digit, address1, address2, mail_address, wechat, cra_sole_proprietor, cra_hst_report, cra_payroll, cra_withhold_tax, cra_wsib, cra_other, oversea_asset_t1135, 
        oversea_corp_t1134, tslip, tax_personal_info, specific_info, engage_account, engage_leading, note, contact_corp, director_corp, sharehold_corp, spouse, parent, child, timemark)
        db.session.add(my_data)
        db.session.commit()
    flash("Import Successfully")
    msg = 'successful'
    return msg

# Fix Indiv address data - remove '-'
def fix_address():
    for i in range (1, 370):
        my_data = Individual.query.get(i)
        da_old = str(my_data.address1)
        da_new = da_old.strip('-').replace(' , ,  , ', '')
        my_data.address1 = da_new
        print('add: ',da_old, '  new  ', da_new)
        db.session.commit()
    
    msg = 'successful'
    return msg

# Database migration
from sqlalchemy import inspect, MetaData, create_engine
def db_export_excel():
    #  (Valid SQLite URL forms are:
    #  sqlite:///:memory: (or, sqlite://)
    #  sqlite:///relative/path/to/file.db
    #  sqlite:////absolute/path/to/file.db )
    # Absolute path of database
    # engine = create_engine('sqlite:////Users/Ritchie/Documents/financial-computing-app/app/databases/users.db')
    # relative path of database
    engine = create_engine('sqlite:///app/databases/users.db')

    # MetaData()方法 ----------------------
    # metadata = MetaData()
    # metadata.reflect(engine)
    # print('metadata: %s' % metadata)
    # for tbl in metadata.tables.values():
    #     print("table %s" % tbl.name)
    # -------------------------------------

    # Inspect()方法 ----------------------
    # inspector = inspect(engine)
    # schemas = inspector.get_schema_names()

    # for schema in schemas:
    #     print("schema: %s" % schema)
    #     for table_name in inspector.get_table_names(schema=schema):
    #         print('table_name %s' % table_name)
    #         for column in inspector.get_columns(table_name, schema=schema):
    #             print("Column: %s" % column)
    #             # 输出的是一个dictionary >>> Column: {'name': 'jobid1', 'type': INTEGER(), 'nullable': True, 'default': None, 'autoincrement': 'auto', 'primary_key': 0}
    #             print("Column: %s" % column['name'])
    #             # 输出dictionary中的键'name'的值 >>> Column: jobid1
    #         print('\n\n\')
    # -------------------------------------

    # python connect database (working fine) ------
    import sqlite3
    con = sqlite3.connect('/Users/Ritchie/Documents/financial-computing-app/app/databases/users.db')
    cursor = con.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    print(cursor.fetchall())
    # >>> [('alembic_version',), ('account',), ('learning_data_table',), ('tbl_activity_code',), ('mulform',), ('timesheettempdata',), ('tbl_corpname_number',), ('tbl_Contacts',), ('tbl_Directors',), ('tbl_Shareholders',), ('tbl_jobType',), ('tbl_Task',), ('num1',), ('num2',), ('tt',), ('tbl_Staff',), ('tbl_Corporation_report',), ('tbl_Individual',), ('tbl_Timesheet',), ('tbl_log',), ('tbl_Corporation',), ('users',)]
    # ------------------------------------------------